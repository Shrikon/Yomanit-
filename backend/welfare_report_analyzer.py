"""
welfare_report_analyzer.py – דוח PDF חודשי לגזבר מקובץ Excel רווחה (תמר)

Usage:
    python welfare_report_analyzer.py <path_to_excel> <output_pdf_path>

מבנה הקובץ (גיליון 'דוח התחשבנות'):
    שורה 4 עמודה 14: שם החודש | שורה 5 עמודה 14: שם הרשות
    שורה 2 עמודה 19: אחוז מימון מצטבר
    עמודה 24: שם סעיף | עמודה 22: מסלול תשלום
    עמודה 7: הקצבה שנתית | עמודה 6: הוצאה מצטברת
    עמודה 4: יתרת הקצבה | עמודה 2: חיוב בחודש זה
    סמל סעיף: נשלף מגיליון 'גרסה להדפסה' עמודה 13 (שורות כותרת סעיף)
    שורות סיכום: col 22 == ' ' (whitespace), שורות משנה: col 22 == 'תשלומי ממשלה' וכו'
"""

import re
import sys
import datetime
import tempfile
import urllib.request
from pathlib import Path
from dataclasses import dataclass
from typing import List, Dict, Optional

import openpyxl
from bidi.algorithm import get_display
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# ── Font ──────────────────────────────────────────────────────────────
_FONT_ZIP_URL = "https://github.com/notofonts/hebrew/releases/download/NotoSansHebrew-v2.004/NotoSansHebrew-v2.004.zip"
_FONT_ZIP_ENTRY = "NotoSansHebrew/full/ttf/NotoSansHebrew-Regular.ttf"
FONT_PATH = Path(tempfile.gettempdir()) / "NotoSansHebrew-Full-Regular.ttf"
FONT_NAME = "NotoSansHebrew"


def _ensure_font():
    """Download Noto Sans Hebrew (full, with digits) to temp dir if not cached."""
    if FONT_PATH.exists() and FONT_PATH.stat().st_size > 50000:
        return
    import zipfile, io
    data = urllib.request.urlopen(_FONT_ZIP_URL).read()
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        with zf.open(_FONT_ZIP_ENTRY) as src, open(FONT_PATH, "wb") as dst:
            dst.write(src.read())

# Header patterns for dynamic column detection (searched in header row)
_HEADER_MAP = {
    "charge":   ["חיוב בחודש זה"],
    "balance":  ["יתרת הקצבה כספית"],
    "expense":  ["הוצאות להתחשבנות מצטברת"],
    "budget":   ["הקצבה שנתית בשקלים", "הקצבה שנתית"],
    "budget_fallback": ["חלק המשרד לפי הסיווג"],
    "units_actual":    ["יחידות ביצוע מצטבר"],
    "units_annual":    ["הקצבה ביחידות שנתי"],
    "classification":  ["אופן סיווג להתחשבנות"],
    "maslul":   ["מסלול תשלום"],
    "name":     ["שם סעיף"],
    "semel":    ["סמל הסעיף", "סמל סעיף"],
}


def _detect_columns(header_row) -> dict:
    """Scan header row and return {field_name: col_index} mapping."""
    cols = {}
    for i, cell in enumerate(header_row):
        val = _safe_str(cell.value)
        if not val:
            continue
        for field, patterns in _HEADER_MAP.items():
            if field in cols:
                continue
            for pat in patterns:
                if pat in val:
                    cols[field] = i
                    break
    return cols


MONTHS_HE = {
    'ינואר': 1, 'פברואר': 2, 'מרץ': 3, 'אפריל': 4,
    'מאי': 5, 'יוני': 6, 'יולי': 7, 'אוגוסט': 8,
    'ספטמבר': 9, 'אוקטובר': 10, 'נובמבר': 11, 'דצמבר': 12,
}


@dataclass
class SectionRow:
    name: str
    semel: str
    budget_annual: float
    expense_cumulative: float
    balance: float
    charge_month: float
    utilization_pct: float
    budget_proportional: float  # הקצבה שנתית × (חודש / 12)
    difference: float           # הוצאה מצטברת − תקציב יחסי


@dataclass
class ReportData:
    municipality: str
    month: str
    month_number: int
    funding_pct: str
    total_budget: float
    total_expense: float
    active_count: int
    rows: List[SectionRow]


# ── Excel Reading ─────────────────────────────────────────────────────

def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _cell(all_rows, row_idx: int, col_idx: int):
    """Get cell value by 0-based row/col index."""
    if row_idx < len(all_rows) and col_idx < len(all_rows[row_idx]):
        return all_rows[row_idx][col_idx].value
    return None


def _extract_semel_from_text(text: str) -> Optional[str]:
    """Extract 6-digit semel from section header like '230090242410 שם סעיף'."""
    m = re.search(r'(\d{6,})', text.strip())
    if m:
        digits = m.group(1)
        # Last 6 digits for long codes, full number for shorter
        return digits[-6:] if len(digits) >= 12 else digits
    return None


def _build_semel_map(wb: openpyxl.Workbook) -> Dict[str, str]:
    """Build name→semel map from sheet 'גרסה להדפסה' section headers."""
    semel_map: Dict[str, str] = {}
    if "גרסה להדפסה" not in wb.sheetnames:
        return semel_map

    ws = wb["גרסה להדפסה"]
    for row in ws.rows:
        # Section header rows: only col 13 has text, rest are empty/None
        if len(row) <= 13:
            continue
        val = row[13].value
        if not val or not isinstance(val, str):
            continue
        val = val.strip()
        # Section header pattern: digits followed by Hebrew text
        m = re.match(r'(\d+)\s+(.+)', val)
        if m:
            semel = _extract_semel_from_text(m.group(1))
            name = m.group(2).strip()
            if semel and name:
                semel_map[name] = semel

    return semel_map


def parse_excel(path: str) -> ReportData:
    """קרא קובץ Excel של דוח התחשבנות והחזר ReportData."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Build semel lookup from sheet 2
    semel_map = _build_semel_map(wb)

    # Open sheet 1
    if "דוח התחשבנות" in wb.sheetnames:
        ws = wb["דוח התחשבנות"]
    elif "גרסה להדפסה" in wb.sheetnames:
        ws = wb["גרסה להדפסה"]
    else:
        raise ValueError(f"גיליון לא נמצא. קיימים: {wb.sheetnames}")

    all_rows = list(ws.rows)

    # If too many None values, retry without data_only (formula cells without cache)
    sample_none = sum(1 for r in all_rows[6:min(12, len(all_rows))] for c in r[:8] if c.value is None)
    if sample_none > 30:
        wb.close()
        wb = openpyxl.load_workbook(path, data_only=False)
        if "דוח התחשבנות" in wb.sheetnames:
            ws = wb["דוח התחשבנות"]
        else:
            ws = wb["גרסה להדפסה"]
        all_rows = list(ws.rows)

    # Metadata — find dynamically by scanning label cells
    def _find_meta(rows, label_text, label_row_hint=None):
        """Find value cell next to a label like 'רשות: ' or 'לחודש: '."""
        search_rows = [rows[label_row_hint]] if label_row_hint and label_row_hint < len(rows) else rows[:8]
        for row in search_rows:
            for i, c in enumerate(row):
                if c.value and label_text in _safe_str(c.value):
                    # Value is in the cell(s) to the LEFT (RTL layout)
                    for j in range(i - 1, -1, -1):
                        val = _safe_str(row[j].value)
                        if val:
                            return val
        return ""

    municipality = _find_meta(all_rows, "רשות:")
    month = _find_meta(all_rows, "לחודש:")
    funding_pct = _find_meta(all_rows, "מימון מצטבר")

    # Resolve month number from Hebrew name
    month_number = MONTHS_HE.get(month, 0)
    if month_number == 0:
        # Try matching partial name
        for name_he, num in MONTHS_HE.items():
            if name_he in month or month in name_he:
                month_number = num
                break

    # Find header row containing 'חיוב בחודש זה'
    header_row_idx = None
    for i in range(min(15, len(all_rows))):
        for c in all_rows[i]:
            if c.value and "חיוב בחודש זה" in _safe_str(c.value):
                header_row_idx = i
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        header_row_idx = 8  # fallback

    # Detect columns dynamically from header row
    col = _detect_columns(all_rows[header_row_idx])

    c_charge = col.get("charge")
    c_balance = col.get("balance")
    c_expense = col.get("expense")
    c_budget = col.get("budget")
    c_budget_fb = col.get("budget_fallback")
    c_units_actual = col.get("units_actual")
    c_units_annual = col.get("units_annual")
    c_classification = col.get("classification")
    c_maslul = col.get("maslul")
    c_name = col.get("name")
    c_semel = col.get("semel")

    if c_name is None:
        raise ValueError(f"Could not find 'שם סעיף' column in header row {header_row_idx}")

    def _rv(row, idx):
        """Read value from row by column index (None-safe)."""
        if idx is not None and idx < len(row):
            return row[idx].value
        return None

    # Parse data — take only summary rows (maslul is whitespace or empty)
    seen_names: dict = {}  # name → SectionRow (deduplicate)
    for i in range(header_row_idx + 1, len(all_rows)):
        row = all_rows[i]

        name = _safe_str(_rv(row, c_name))
        if not name:
            continue
        if 'סה"כ' in name or "סה''כ" in name:
            continue

        # Filter: only summary rows (maslul is whitespace or empty)
        maslul = _safe_str(_rv(row, c_maslul))
        if maslul and maslul not in (" ", ""):
            continue

        budget_raw = _rv(row, c_budget)
        budget_fb = _rv(row, c_budget_fb)
        # Use fallback when primary is None or 0 but fallback has a real value
        budget = _safe_float(budget_raw)
        if budget == 0 and _safe_float(budget_fb) != 0:
            budget = _safe_float(budget_fb)

        # Quantitative sections: col7=None and classification='כמותי'
        classification = _safe_str(_rv(row, c_classification))
        if budget_raw is None and "כמותי" in classification:
            expense_val = _safe_float(_rv(row, c_expense))
            units_actual = _safe_float(_rv(row, c_units_actual))
            units_annual = _safe_float(_rv(row, c_units_annual))
            if expense_val > 0 and units_actual > 0:
                tariff = expense_val / units_actual
                budget = round(units_annual * tariff, 2)
            else:
                continue  # skip section with no data

        expense = _safe_float(_rv(row, c_expense))
        balance = _safe_float(_rv(row, c_balance))
        charge = _safe_float(_rv(row, c_charge))

        if budget == 0 and expense == 0 and balance == 0:
            continue

        utilization = (expense / budget * 100) if budget != 0 else 0.0
        semel = _safe_str(_rv(row, c_semel))
        if not semel:
            semel = semel_map.get(name, "")
        proportional = budget * (month_number / 12) if month_number else 0.0
        diff = expense - proportional

        seen_names[name] = SectionRow(
            name=name,
            semel=semel,
            budget_annual=budget,
            expense_cumulative=expense,
            balance=balance,
            charge_month=charge,
            utilization_pct=round(utilization, 1),
            budget_proportional=round(proportional, 2),
            difference=round(diff, 2),
        )

    wb.close()

    rows = list(seen_names.values())
    total_budget = sum(r.budget_annual for r in rows)
    total_expense = sum(r.expense_cumulative for r in rows)

    return ReportData(
        municipality=municipality,
        month=month,
        month_number=month_number,
        funding_pct=funding_pct,
        total_budget=total_budget,
        total_expense=total_expense,
        active_count=len(rows),
        rows=rows,
    )


# ── PDF Generation (Canvas-based for reliable Hebrew rendering) ───────

def _register_font():
    _ensure_font()
    pdfmetrics.registerFont(TTFont(FONT_NAME, str(FONT_PATH)))


def _fmt_num(val: float) -> str:
    if val == int(val):
        return f"{int(val):,}"
    return f"{val:,.2f}"


def _bidi(text: str) -> str:
    """Apply BiDi algorithm to convert logical Hebrew to visual order for PDF."""
    return get_display(str(text))


# Colors
CLR_HEADER_BG = (0.17, 0.24, 0.31)   # #2c3e50
CLR_HEADING = (0.10, 0.32, 0.46)      # #1a5276
CLR_ALT_ROW = (0.92, 0.95, 0.97)      # #eaf2f8
CLR_GRID = (0.6, 0.6, 0.6)


class _PdfWriter:
    """Canvas-based PDF writer with RTL table support."""

    def __init__(self, path: str):
        self.c = canvas.Canvas(path, pagesize=A4)
        self.page_w, self.page_h = A4
        self.margin = 15 * mm
        self.y = self.page_h - self.margin
        self.content_w = self.page_w - 2 * self.margin

    def _check_space(self, needed: float):
        if self.y - needed < self.margin:
            self.c.showPage()
            self.y = self.page_h - self.margin

    def draw_title(self, text: str):
        self._check_space(25)
        self.c.setFont(FONT_NAME, 16)
        self.c.setFillColorRGB(0, 0, 0)
        self.c.drawCentredString(self.page_w / 2, self.y, _bidi(text))
        self.y -= 25

    def draw_heading(self, text: str):
        self._check_space(50)
        self.y -= 10
        self.c.setFont(FONT_NAME, 12)
        self.c.setFillColorRGB(*CLR_HEADING)
        self.c.drawRightString(self.page_w - self.margin, self.y, _bidi(text))
        self.c.setFillColorRGB(0, 0, 0)
        self.y -= 30

    def draw_text(self, text: str, size: int = 9):
        self._check_space(14)
        self.c.setFont(FONT_NAME, size)
        self.c.setFillColorRGB(0, 0, 0)
        self.c.drawRightString(self.page_w - self.margin, self.y, _bidi(text))
        self.y -= 14

    def draw_kv_table(self, rows: List[List[str]]):
        """Draw a simple 2-column key-value table (RTL)."""
        row_h = 18
        total_h = len(rows) * row_h
        self._check_space(total_h + 5)

        x_right = self.page_w - self.margin
        col1_w = 55 * mm  # label column (right)
        col2_w = self.content_w - col1_w  # value column (left)

        for i, (label, value) in enumerate(rows):
            row_y = self.y - (i * row_h)
            # Alternating background
            if i % 2 == 1:
                self.c.setFillColorRGB(*CLR_ALT_ROW)
                self.c.rect(self.margin, row_y - 4, self.content_w, row_h, fill=1, stroke=0)
            # Grid lines
            self.c.setStrokeColorRGB(*CLR_GRID)
            self.c.setLineWidth(0.3)
            self.c.line(self.margin, row_y - 4, x_right, row_y - 4)
            # Label (right-aligned, right column)
            self.c.setFillColorRGB(0, 0, 0)
            self.c.setFont(FONT_NAME, 10)
            self.c.drawRightString(x_right - 3, row_y + 2, _bidi(label))
            # Value (right-aligned in left column)
            self.c.drawRightString(x_right - col1_w - 3, row_y + 2, _bidi(value))

        self.y -= total_h + 8

    def _truncate(self, text: str, max_w: float, font_size: float) -> str:
        """Truncate text with '...' if it exceeds max_w."""
        if self.c.stringWidth(text, FONT_NAME, font_size) <= max_w:
            return text
        while len(text) > 1:
            text = text[:-1]
            if self.c.stringWidth(text + "...", FONT_NAME, font_size) <= max_w:
                return text + "..."
        return "..."

    def draw_data_table(self, headers: List[str], rows: List[List[str]]):
        """Draw a data table with auto-sized columns. Columns in RTL order."""
        ROW_H = 20
        HEADER_H = 22
        CELL_PAD = 6          # horizontal padding inside cell
        HEADER_FONT = 7.5
        DATA_FONT = 7.5

        # Reverse for RTL display
        headers = list(reversed(headers))
        rows = [list(reversed(r)) for r in rows]
        num_cols = len(headers)

        # ── Auto-size columns ──
        # Measure widest content per column (header + all data rows)
        min_widths = []
        for col_idx in range(num_cols):
            hw = self.c.stringWidth(_bidi(headers[col_idx]),
                                    FONT_NAME, HEADER_FONT)
            max_data = 0
            for row in rows:
                if col_idx < len(row):
                    dw = self.c.stringWidth(_bidi(row[col_idx]),
                                            FONT_NAME, DATA_FONT)
                    if dw > max_data:
                        max_data = dw
            min_widths.append(max(hw, max_data) + CELL_PAD)

        total_natural = sum(min_widths)

        if total_natural <= self.content_w:
            # Distribute remaining space proportionally
            extra = self.content_w - total_natural
            col_widths_final = [w + extra * (w / total_natural)
                                for w in min_widths]
        else:
            # Shrink proportionally — longest columns shrink most
            col_widths_final = [w * (self.content_w / total_natural)
                                for w in min_widths]

        x_left = self.margin
        total_w = sum(col_widths_final)

        def _draw_row(cells, y, h, is_header=False, bg=None):
            # Background
            if is_header:
                self.c.setFillColorRGB(*CLR_HEADER_BG)
                self.c.rect(x_left, y - 4, total_w, h, fill=1, stroke=0)
            elif bg:
                self.c.setFillColorRGB(*bg)
                self.c.rect(x_left, y - 4, total_w, h, fill=1, stroke=0)

            fs = HEADER_FONT if is_header else DATA_FONT
            if is_header:
                self.c.setFillColorRGB(1, 1, 1)
            else:
                self.c.setFillColorRGB(0, 0, 0)
            self.c.setFont(FONT_NAME, fs)

            x = x_left
            for j, cell_text in enumerate(cells):
                cw = col_widths_final[j] if j < len(col_widths_final) \
                    else col_widths_final[-1]
                usable = cw - CELL_PAD
                display_text = _bidi(self._truncate(cell_text, usable, fs))
                text_w = self.c.stringWidth(display_text, FONT_NAME, fs)
                text_x = x + (cw - text_w) / 2  # center in cell
                self.c.drawString(text_x, y + 4, display_text)
                x += cw

            # Grid
            self.c.setStrokeColorRGB(*CLR_GRID)
            self.c.setLineWidth(0.5)
            # Horizontal line at bottom of row
            self.c.line(x_left, y - 4, x_left + total_w, y - 4)
            # Vertical lines
            x = x_left
            for cw in col_widths_final:
                self.c.line(x, y - 4, x, y - 4 + h)
                x += cw
            self.c.line(x, y - 4, x, y - 4 + h)

        # ── Draw header ──
        self._check_space(HEADER_H + ROW_H + 10)
        _draw_row(headers, self.y, HEADER_H, is_header=True)
        self.y -= HEADER_H

        # ── Draw data rows ──
        for i, row in enumerate(rows):
            self._check_space(ROW_H + 5)
            bg = CLR_ALT_ROW if i % 2 == 1 else None
            _draw_row(row, self.y, ROW_H, bg=bg)
            self.y -= ROW_H

        # Top border of last row (bottom of table)
        self.c.line(x_left, self.y + ROW_H - 4, x_left + total_w,
                    self.y + ROW_H - 4)
        self.y -= 8

    def spacer(self, h: float = 8):
        self.y -= h

    def save(self):
        self.c.save()


def generate_pdf(report: ReportData, output_path: str) -> str:
    """Generate the PDF report. Returns the output path."""
    _register_font()
    pdf = _PdfWriter(output_path)
    now = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")

    # ── Title ──
    pdf.draw_title(f"דוח תקצוב והתחשבנות — {report.municipality}")
    pdf.spacer(6)

    # ── Section 1: Executive Summary ──
    pdf.draw_heading("סיכום מנהלים")
    total_proportional = sum(r.budget_proportional for r in report.rows)
    total_diff = sum(r.difference for r in report.rows)

    pdf.draw_kv_table([
        ["רשות", report.municipality],
        ["חודש דיווח", f"{report.month} ({report.month_number}/12)"],
        ["אחוז מימון מצטבר", report.funding_pct],
        ['סה"כ הקצבה שנתית', _fmt_num(report.total_budget)],
        ["תקציב יחסי לתקופה", _fmt_num(total_proportional)],
        ['סה"כ הוצאה מצטברת', _fmt_num(report.total_expense)],
        ["הפרש מצטבר", _fmt_num(total_diff)],
        ["סעיפים פעילים", str(report.active_count)],
        ["תאריך הפקה", now],
    ])
    pdf.spacer(10)

    # ── Section 2: Underutilized (balance > 20% of annual) ──
    pdf.draw_heading("סעיפים עם תקציב לא מנוצל (יתרה > 20% מהקצבה)")

    underutilized = [
        r for r in report.rows
        if r.budget_annual > 0 and r.balance > r.budget_annual * 0.2
    ]
    underutilized.sort(key=lambda r: r.balance, reverse=True)

    if underutilized:
        headers = ["שם סעיף", "סמל", "הקצבה שנתית", "תקציב יחסי",
                   "הוצאה מצטברת", "הפרש", "יתרה", "% ניצול"]
        data = [
            [r.name, r.semel, _fmt_num(r.budget_annual),
             _fmt_num(r.budget_proportional), _fmt_num(r.expense_cumulative),
             _fmt_num(r.difference), _fmt_num(r.balance),
             f"{r.utilization_pct}%"]
            for r in underutilized
        ]
        pdf.draw_data_table(headers, data)
    else:
        pdf.draw_text(".אין סעיפים עם יתרה מעל 20%")

    pdf.spacer(10)

    # ── Section 3: Overbudget (expense > budget) ──
    pdf.draw_heading("סעיפים עם חריגה (הוצאה > הקצבה)")

    overbudget = [
        r for r in report.rows
        if r.budget_annual > 0 and r.expense_cumulative > r.budget_annual
    ]
    overbudget.sort(
        key=lambda r: r.expense_cumulative - r.budget_annual, reverse=True
    )

    if overbudget:
        headers = ["שם סעיף", "סמל", "הקצבה שנתית", "תקציב יחסי",
                   "הוצאה מצטברת", "הפרש", "חריגה שנתית"]
        data = [
            [r.name, r.semel, _fmt_num(r.budget_annual),
             _fmt_num(r.budget_proportional), _fmt_num(r.expense_cumulative),
             _fmt_num(r.difference),
             _fmt_num(r.expense_cumulative - r.budget_annual)]
            for r in overbudget
        ]
        pdf.draw_data_table(headers, data)
    else:
        pdf.draw_text(".אין סעיפים עם חריגה תקציבית")

    pdf.save()
    return output_path


# ── CLI ───────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) != 3:
        print("Usage: python welfare_report_analyzer.py <excel_path> <output_pdf>")
        sys.exit(1)

    excel_path = sys.argv[1]
    output_pdf = sys.argv[2]

    report = parse_excel(excel_path)

    print(f"[ANALYZER] רשות: {report.municipality}")
    print(f"[ANALYZER] חודש: {report.month}")
    print(f"[ANALYZER] מימון מצטבר: {report.funding_pct}")
    print(f"[ANALYZER] סעיפים פעילים: {report.active_count}")
    print(f'[ANALYZER] סה"כ הקצבה: {_fmt_num(report.total_budget)}')
    print(f'[ANALYZER] סה"כ הוצאה: {_fmt_num(report.total_expense)}')

    underutilized = [
        r for r in report.rows
        if r.budget_annual > 0 and r.balance > r.budget_annual * 0.2
    ]
    overbudget = [
        r for r in report.rows
        if r.budget_annual > 0 and r.expense_cumulative > r.budget_annual
    ]
    print(f"[ANALYZER] סעיפים לא מנוצלים: {len(underutilized)}")
    print(f"[ANALYZER] סעיפים בחריגה: {len(overbudget)}")

    generate_pdf(report, output_pdf)
    print(f"[ANALYZER] PDF נוצר: {output_pdf}")


if __name__ == "__main__":
    main()

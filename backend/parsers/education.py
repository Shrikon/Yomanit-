"""
education.py – מאחד קבצי חינוך (CSV, קידוד windows-1255) ומייצר Excel בקרה.

Usage:
    python -m parsers.education <folder_path> <output.xlsx>

קבצים מותרים: GY, HASAOT, MUTAVIM, SHARATIM, SHEFI, YADANIIM,
              MUCARIM, MOADON, ICHLUS, AZAROLIM, SACAL
קבצים אסורים: HASNET, HASMASLULIM, MISROT, MISROTGY, SACALCHARIGIM
"""

import csv
import sys
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Allowed / blocked files ───────────────────────────────────────────

ALLOWED_FILES = {
    "GY", "HASAOT", "MUTAVIM", "SHARATIM", "SHEFI", "YADANIIM",
    "MUCARIM", "MOADON", "ICHLUS", "AZAROLIM", "SACAL",
}
BLOCKED_FILES = {
    "HASNET", "HASMASLULIM", "MISROT", "MISROTGY", "SACALCHARIGIM",
}

CHESHBONIT_NAME = "CHESHBONIT"


# ── Helpers ───────────────────────────────────────────────────────────

def _file_tag(filename: str) -> str:
    """Extract tag from filename (e.g. 'GY_2026.csv' → 'GY')."""
    stem = Path(filename).stem.upper()
    for tag in sorted(ALLOWED_FILES | BLOCKED_FILES | {CHESHBONIT_NAME}, key=len, reverse=True):
        if stem.startswith(tag):
            return tag
    return stem


def _read_csv(path: Path) -> list[dict]:
    """Read a windows-1255 CSV, strip BOM, return list of row dicts."""
    raw = path.read_bytes()
    if raw[:3] == b"\xef\xbb\xbf":
        raw = raw[3:]
    text = raw.decode("windows-1255", errors="replace")
    lines = text.splitlines()
    if not lines:
        return []
    reader = csv.DictReader(lines)
    return list(reader)


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    val = str(val).strip().replace(",", "")
    if not val:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _make_key(row: dict) -> str | None:
    code = (row.get("קוד נושא") or "").strip()
    month = (row.get("חודש תחולה") or "").strip()
    if not code:
        return None
    return f"{code}|{month}"


# ── Core logic ────────────────────────────────────────────────────────

def aggregate_sources(folder: Path) -> dict[str, float]:
    """Read all allowed CSV files and aggregate amounts by key."""
    totals: dict[str, float] = defaultdict(float)
    files_read = 0

    for csv_path in sorted(folder.glob("*.csv")):
        tag = _file_tag(csv_path.name)

        if tag in BLOCKED_FILES:
            print(f"  [SKIP] {csv_path.name} (blocked)")
            continue
        if tag == CHESHBONIT_NAME:
            continue
        if tag not in ALLOWED_FILES:
            print(f"  [SKIP] {csv_path.name} (unknown)")
            continue

        rows = _read_csv(csv_path)
        amount_col = "סכום מחושב" if tag == "YADANIIM" else "הפרש מחושב"
        count = 0
        for row in rows:
            key = _make_key(row)
            if key is None:
                continue
            totals[key] += _safe_float(row.get(amount_col))
            count += 1

        files_read += 1
        print(f"  [OK]   {csv_path.name} ({tag}): {count} rows")

    print(f"  Source files read: {files_read}, unique keys: {len(totals)}")
    return dict(totals)


def read_cheshbonit(folder: Path) -> dict[str, float]:
    """Read CHESHBONIT CSV and return amounts by key."""
    totals: dict[str, float] = defaultdict(float)

    for csv_path in sorted(folder.glob("*.csv")):
        tag = _file_tag(csv_path.name)
        if tag != CHESHBONIT_NAME:
            continue

        rows = _read_csv(csv_path)
        count = 0
        for row in rows:
            key = _make_key(row)
            if key is None:
                continue
            totals[key] += _safe_float(row.get("יתרת ביצוע החודש"))
            count += 1

        print(f"  [CH]   {csv_path.name}: {count} rows")
        break

    return dict(totals)


def build_control(
    source: dict[str, float], cheshbonit: dict[str, float]
) -> list[dict]:
    """Build control rows comparing source vs cheshbonit."""
    all_keys = sorted(set(source) | set(cheshbonit))
    rows = []
    for key in all_keys:
        code, month = key.split("|", 1) if "|" in key else (key, "")
        src = source.get(key, 0.0)
        ch = cheshbonit.get(key, 0.0)
        diff = round(src - ch, 2)
        final = src if diff == 0 else ch
        has_adj = diff != 0
        if diff == 0:
            status = "MATCH"
        elif src == 0:
            status = "CH_ONLY"
        elif ch == 0:
            status = "SRC_ONLY"
        else:
            status = "DIFF"
        rows.append({
            "code": code,
            "month": month,
            "source": round(src, 2),
            "ch": round(ch, 2),
            "diff": diff,
            "final": round(final, 2),
            "has_adj": has_adj,
            "status": status,
        })
    return rows


# ── Excel output ──────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
_ALT_FILL = PatternFill("solid", fgColor="EAF2F8")
_DIFF_FILL = PatternFill("solid", fgColor="FADBD8")
_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)


def write_excel(control: list[dict], output_path: str):
    wb = openpyxl.Workbook()

    # ── Sheet 1: RESULT ──
    ws1 = wb.active
    ws1.title = "RESULT"
    ws1.sheet_view.rightToLeft = True

    result_headers = ["קוד נושא", "חודש תחולה", "סכום סופי"]
    for c, h in enumerate(result_headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for i, row in enumerate(control, 2):
        ws1.cell(row=i, column=1, value=row["code"]).border = _BORDER
        ws1.cell(row=i, column=2, value=row["month"]).border = _BORDER
        ws1.cell(row=i, column=3, value=row["final"]).border = _BORDER
        if i % 2 == 0:
            for c in range(1, 4):
                ws1.cell(row=i, column=c).fill = _ALT_FILL

    for col in [1, 2, 3]:
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    # ── Sheet 2: CONTROL ──
    ws2 = wb.create_sheet("CONTROL")
    ws2.sheet_view.rightToLeft = True

    ctrl_headers = [
        "קוד נושא", "חודש תחולה", "SourceAmount", "CHAmount",
        "Diff", "HasAdjustment", "MatchStatus",
    ]
    for c, h in enumerate(ctrl_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for i, row in enumerate(control, 2):
        ws2.cell(row=i, column=1, value=row["code"]).border = _BORDER
        ws2.cell(row=i, column=2, value=row["month"]).border = _BORDER
        ws2.cell(row=i, column=3, value=row["source"]).border = _BORDER
        ws2.cell(row=i, column=4, value=row["ch"]).border = _BORDER
        cell_diff = ws2.cell(row=i, column=5, value=row["diff"])
        cell_diff.border = _BORDER
        if row["diff"] != 0:
            cell_diff.fill = _DIFF_FILL
        ws2.cell(row=i, column=6, value="V" if row["has_adj"] else "").border = _BORDER
        ws2.cell(row=i, column=7, value=row["status"]).border = _BORDER
        if i % 2 == 0:
            for c in range(1, 8):
                cell = ws2.cell(row=i, column=c)
                if not cell.fill or cell.fill.fgColor.rgb == "00000000":
                    cell.fill = _ALT_FILL

    for col in range(1, 8):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    wb.save(output_path)


# ── CLI ───────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) != 3:
        print("Usage: python -m parsers.education <folder_path> <output.xlsx>")
        sys.exit(1)

    folder = Path(sys.argv[1])
    output = sys.argv[2]

    if not folder.is_dir():
        print(f"Error: {folder} is not a directory")
        sys.exit(1)

    print(f"Reading CSVs from: {folder}")
    source = aggregate_sources(folder)
    cheshbonit = read_cheshbonit(folder)
    control = build_control(source, cheshbonit)
    write_excel(control, output)

    total_src = sum(r["source"] for r in control)
    total_ch = sum(r["ch"] for r in control)
    total_final = sum(r["final"] for r in control)

    print(f"\nSummary:")
    print(f"  Total Source:  {total_src:,.2f}")
    print(f"  Total CH:      {total_ch:,.2f}")
    print(f"  Total Final:   {total_final:,.2f}")
    print(f"  Rows: {len(control)} ({sum(1 for r in control if r['status']=='MATCH')} match, "
          f"{sum(1 for r in control if r['status']=='DIFF')} diff)")
    print(f"  Output: {output}")


if __name__ == "__main__":
    main()

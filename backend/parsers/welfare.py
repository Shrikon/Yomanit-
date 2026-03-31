# parsers/welfare.py РђЊ ОцОеОАОе ОДОЋОЉОЦ ОеОЋОЋОЌОћ ОњОЋОюОъОЎ (ОфОъОе)
# ОюОЋОњОЎОДОћ:
# ОЌОЋОЉОћ = ОеОД ОЕОЋОеОЋОф ОфОЕОюОЋОъОЎ ОъОъОЕОюОћ (ОАОЏОЋОЮ ОЏОю ОћОЕОЋОеОЋОф ОюОљОЋОфОЋ ОАОбОЎОБ) + ОЎОЕ debit ОЉ-INDEX
# ОќОЏОЋОф = ОќОЎОЏОЋОЎ/ОЌОЎОЋОЉ ОЉОЌОЋОЊОЕ ОќОћ ОъОћОЕОЋОеОћ ОћОеОЎОДОћ + ОЎОЕ credit ОЉ-INDEX

import io, re, datetime
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from typing import List, Dict, Any, Tuple
from openpyxl import load_workbook


class WelfareParserError(Exception):
    pass


# РћђРћђ INDEX: ОАОъОю ОАОбОЎОБ Рєњ ОДОЋОЊОЎ ОЌОЕОЉОЋОЪ РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
# debit  = ОфОДОдОЎОЉ ОЌОЋОЉОћ (ОеОД ОюОАОбОЎОцОЎОЮ ОбОЮ ОфОЕОюОЋОъОЎ ОъОъОЕОюОћ)
# credit = ОфОДОдОЎОЉ ОќОЏОЋОф
WELFARE_INDEX = {
    # РћђРћђ ОЌОЎОеОЋОЮ РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    "120211": {"debit": "1849999783", "credit": "1342212930"},
    "120214": {"credit": "1342212930"},
    "120217": {"debit": "1842203840", "credit": "1342203930"},
    "120218": {"credit": "1342212930"},
    # РћђРћђ ОљОќОеОЌОЎОЮ ОЋОфОЎОДОЎОЮ ОЋОаОЎОдОЋОюОЎ ОЕОЋОљОћ РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    "242410": {"debit": "1844301840", "credit": "1344301930"},
    "243410": {"debit": "1844409750", "credit": "1344409930"},
    "243415": {"debit": "1844419840", "credit": "1344416930"},
    "243417": {"debit": "1844414840", "credit": "1344414930"},
    "243418": {"debit": "1844408840", "credit": "1344408930"},
    "243419": {"debit": "1844501840", "credit": "1344501930"},
    "243420": {"credit": "1344409930"},
    "243430": {"debit": "1844414840", "credit": "1344414930"},
    "243438": {"debit": "1844408840", "credit": "1344408930"},
    # РћђРћђ ОЏОЋОЌ ОљОЊОЮ РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    "513410": {"credit": "1341000930"},
    "513411": {"credit": "1341201930"},
    "513412": {"credit": "1341201930"},
    "513420": {"debit": "1841001840", "credit": "1341000932"},
    "513421": {"credit": "1341004930"},
    "513423": {"credit": "1341004930"},
    "513440": {"credit": "1341003930"},
    "513441": {"credit": "1341001930"},
    # РћђРћђ ОаОЏОЎОЮ ОЋОЕОЎОДОЋОЮ РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    "721010": {"debit": "1846501840", "credit": "1346501930"},
    "721011": {"debit": "1846502840", "credit": "1346502930"},
    "721012": {"debit": "1846100840", "credit": "1346100930"},
    "721020": {"debit": "1845108840", "credit": "1345108930"},
    "721030": {"debit": "1845103840", "credit": "1345103930"},
    "721040": {"debit": "1845110840", "credit": "1345110930"},
    "721050": {"debit": "1845105840", "credit": "1345105930"},
    "721060": {"debit": "1845104840", "credit": "1345104930"},
    "722020": {"debit": "1845203840", "credit": "1345203930"},
    "722021": {"debit": "1845204840", "credit": "1345204930"},
    "722022": {"debit": "1845204840", "credit": "1345204930"},
    "722030": {"debit": "1845213840", "credit": "1345213930"},
    "722041": {"debit": "1845209840", "credit": "1345209930"},
    "722042": {"debit": "1845210840", "credit": "1345210930"},
    "723010": {"debit": "1845303840", "credit": "1345303930"},
    "723011": {"debit": "1845305840", "credit": "1345305930"},
    "723012": {"debit": "1845306840", "credit": "1345306930"},
    "723013": {"debit": "1845308840", "credit": "1345308930"},
    "723014": {"debit": "1845308840", "credit": "1345308930"},
    "723020": {"debit": "1845310840", "credit": "1345310930"},
    "723050": {"debit": "1845325840", "credit": "1345325930"},
    "723051": {"debit": "1845326840", "credit": "1345326930"},
    "723054": {"debit": "1845332840", "credit": "1345332930"},
    "723056": {"debit": "1845334840", "credit": "1345334930"},
    "723210": {"debit": "1845404840", "credit": "1345404930"},
    "723212": {"debit": "1845408840", "credit": "1345408930"},
    "723214": {"debit": "1845410840", "credit": "1345410930"},
    "723215": {"debit": "1845411840", "credit": "1345411930"},
    "723217": {"debit": "1845111840", "credit": "1345111930"},
    "723220": {"debit": "1845414840", "credit": "1345414930"},
    "723221": {"debit": "1845415840", "credit": "1345415930"},
    "723224": {"debit": "1845419840", "credit": "1345419930"},
    "723225": {"debit": "1845420840", "credit": "1345420930"},
    "723670": {"debit": "1845502840", "credit": "1345502930"},
    "723671": {"debit": "1845503840", "credit": "1345503930"},
    "723820": {"debit": "1845601840", "credit": "1345601930"},
    # РћђРћђ ОЎОюОЊОЎОЮ ОЋОъОЕОцОЌОћ РћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђРћђ
    "1038100": {"debit": "1847101840", "credit": "1347101930"},
    "1038400": {"debit": "1847201840", "credit": "1347201930"},
    "1038405": {"debit": "1847202840", "credit": "1347202930"},
    "1038410": {"debit": "1847203840", "credit": "1347203930"},
    "1038411": {"debit": "1847204840", "credit": "1347204930"},
    "1038413": {"debit": "1847206840", "credit": "1347206930"},
    "1039010": {"debit": "1847301840", "credit": "1347301930"},
    "1039320": {"debit": "1847401840", "credit": "1347401930"},
    "1039440": {"debit": "1847501840", "credit": "1347501930"},
    "1039441": {"debit": "1847502840", "credit": "1347502930"},
    "1039448": {"debit": "1847508840", "credit": "1347508930"},
    "1039483": {"debit": "1847511840", "credit": "1347511930"},
    "1039730": {"debit": "1847601840", "credit": "1347601930"},
    "1039733": {"debit": "1847602840", "credit": "1347602930"},
    "1039750": {"debit": "1847604840", "credit": "1347604930"},
    "1039751": {"debit": "1847605840", "credit": "1347605930"},
    "1039770": {"debit": "1847607840", "credit": "1347607930"},
    "1039790": {"debit": "1847609840", "credit": "1347609930"},
    "1039810": {"debit": "1847611840", "credit": "1347611930"},
    "1165600": {"debit": "1848101840", "credit": "1348101930"},
    "1175060": {"debit": "1848201840", "credit": "1348201930"},
    "1175063": {"debit": "1848202840", "credit": "1348202930"},
    "1175320": {"debit": "1847106840", "credit": "1347106930"},
    "1175330": {"debit": "1848301840", "credit": "1348301930"},
    "1175331": {"debit": "1848302840", "credit": "1348302930"},
    "1175332": {"debit": "1848303840", "credit": "1348303930"},
    "1175370": {"debit": "1848401840", "credit": "1348401930"},
}


def _extract_semel(raw: str) -> str:
    """ОъОЌОюОЦ ОъОАОцОе ОАОбОЎОБ ОъОћОфОљ Рђћ ОъОАОЎОе ОаОДОЋОЊОЋОф ОЋОеОЋОЋОЌОЎОЮ."""
    s = str(raw).strip()
    s = s.split('.')[0]
    s = s.replace(' ', '').replace('\u200b', '')
    return s


def _to_decimal(val) -> Decimal:
    try:
        if val is None:
            return Decimal("0")
        if isinstance(val, float) and (val != val):  # NaN check
            return Decimal("0")
        return Decimal(str(val)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, TypeError):
        return Decimal("0")


def parse_welfare(content: bytes, month: int = None, index_map: Dict[str, Dict] = None) -> Dict[str, Any]:
    welfare_index = index_map if index_map is not None else WELFARE_INDEX

    # ОДОеОЎОљОћ ОъОћОЎОеОћ ОЉopenpyxl read_only=True Рђћ ОюОљ ОўОЋОбОЪ DataFrame ОюОќОЎОЏОеОЋОЪ
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise WelfareParserError(f"ОюОљ ОаОЎОфОЪ ОюОДОеОЋОљ ОљОф ОћОДОЋОЉОЦ: {e}")

    sheet_name = None
    for name in wb.sheetnames:
        if 'ОЊОЋОЌ ОћОфОЌОЕОЉОаОЋОф' in name or 'ОњОеОАОћ ОюОћОЊОцОАОћ' in name:
            sheet_name = name
            break
    if not sheet_name:
        wb.close()
        raise WelfareParserError("ОДОЋОЉОЦ ОюОљ ОфОДОаОЎ РђЊ ОЌОАОе sheet 'ОЊОЋОЌ ОћОфОЌОЕОЉОаОЋОф'")

    ws = wb[sheet_name]
    # ОўОбОЋОЪ ОљОф ОЏОю ОћОЕОЋОеОЋОф ОюОеОЕОЎОъОћ ОцОбОЮ ОљОЌОф Рђћ tuple ОДОю ОЎОЋОфОе Оъ-DataFrame
    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()

    def _cell(row, col):
        if col is None or col >= len(row):
            return None
        return row[col]

    def _str(val):
        if val is None:
            return ''
        return str(val).strip()

    def _float(val):
        if val is None:
            return 0.0
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0.0

    # ОЕОЮ ОеОЕОЋОф
    municipality = ""
    for row in rows_raw[:8]:
        for v in row:
            s = _str(v)
            if any(x in s for x in ['Оъ. Ољ', 'ОъОЋОбОдОћ', 'ОбОЎОеОЎОЎОф', 'Оъ.Ољ', 'Оъ.Оъ']):
                municipality = s
                break

    # ОЌОЋОЊОЕ ОЋОЕОаОћ
    period_label = ""
    period_month = month
    period_year  = None
    for row in rows_raw[:8]:
        for v in row:
            s = _str(v)
            m = re.search(r'ОфОЕОюОЋОЮ ОюОЌОЋОЊОЕ\s*(\d+)/(\d{4})', s)
            if m:
                period_month = int(m.group(1))
                period_year  = int(m.group(2))
                period_label = f"{period_month}/{period_year}"
                break

    if period_month is None:
        period_month = 1
    if period_year is None:
        period_year = datetime.datetime.now().year

    # ОъОдОљ ОЕОЋОеОф headers
    header_row_idx = None
    col_semel = col_name = col_maslul = col_total = col_zikuy = None
    summary_mishrad = None
    summary_choz    = None

    for i, row in enumerate(rows_raw[:15]):
        vals = [_str(v) for v in row]
        if 'ОЌОЎОЋОЉ ОЉОЌОЋОЊОЕ ОќОћ' in vals:
            header_row_idx = i
            col_semel  = next((j for j, v in enumerate(vals) if 'ОАОъОю ОћОАОбОЎОБ' in v), None)
            col_name   = next((j for j, v in enumerate(vals) if v == 'ОЕОЮ ОАОбОЎОБ'), None)
            col_maslul = next((j for j, v in enumerate(vals) if 'ОъОАОюОЋОю ОфОЕОюОЋОЮ' in v), None)
            col_total  = next((j for j, v in enumerate(vals) if 'ОАОћ"ОЏ ОћОЋОдОљОћ' in v), None)
            col_zikuy  = next((j for j, v in enumerate(vals) if 'ОќОЎОЏОЋОЎ/ОЌОЎОЋОЉ ОЉОЌОЋОЊОЕ' in v), None)
            break

    if header_row_idx is None:
        raise WelfareParserError("ОюОљ ОаОъОдОљОћ ОЕОЋОеОф ОЏОЋОфОеОЋОф ОЉОДОЋОЉОЦ")
    if col_semel is None:
        raise WelfareParserError("ОбОъОЋОЊОф 'ОАОъОю ОћОАОбОЎОБ' ОюОљ ОаОъОдОљОћ")

    # ОъОдОљ ОАОЎОЏОЋОъОЎОЮ ОъОћОЊОЋОЌ
    KEYWORDS_MISHRAD = ['ОфОЕОюОЋОъОЎ ОъОъОЕОюОћ']
    KEYWORDS_CHOZ    = ['ОЌОЎОЋОЉ/ОќОЎОЏОЋОЎ ОеОЕОЋОф', 'ОќОЎОЏОЋОЎ ОеОЕОЋОф']

    def _find_summary_value(keywords):
        for row in reversed(rows_raw[header_row_idx:]):
            row_text = ' '.join(_str(v) for v in row)
            if any(k in row_text for k in keywords):
                for val in row:
                    if val is None:
                        continue
                    s = str(val).replace(',', '').strip()
                    if re.fullmatch(r'-?\d+(\.\d+)?', s):
                        return abs(float(s))
        return None

    summary_mishrad = _find_summary_value(KEYWORDS_MISHRAD)
    summary_choz    = _find_summary_value(KEYWORDS_CHOZ)

    # ОљОАОЋОБ ОаОфОЋОаОЎОЮ ОюОцОЎ ОАОбОЎОБ
    semel_data: Dict[str, dict] = {}
    EXCLUDE_MASLUL = ['ОћОъОЌОљОЋОф', 'ОЕОўОеОЮ ОаОцОЊОЋ', 'ОъОАОе']

    for row in rows_raw[header_row_idx + 1:]:
        raw_semel = _str(_cell(row, col_semel))
        if not raw_semel or raw_semel == 'nan':
            continue

        semel  = _extract_semel(raw_semel)
        maslul = _str(_cell(row, col_maslul))
        name   = _str(_cell(row, col_name))
        total  = _float(_cell(row, col_total))
        zikuy  = _float(_cell(row, col_zikuy))

        if semel not in semel_data:
            semel_data[semel] = {
                'name':       name or semel,
                'has_ОъОъОЕОюОћ': False,
                'debit_total': 0,
                'zikuy':      0,
            }

        if (maslul and
                'ОеОЕОЋОф' not in maslul and
                'ОЎОюОЊОЎ ОЌОЋОЦ' not in maslul and
                not any(k in maslul for k in EXCLUDE_MASLUL)):
            semel_data[semel]['has_ОъОъОЕОюОћ'] = True
            semel_data[semel]['debit_total'] += total

        if (not maslul) and zikuy != 0:
            semel_data[semel]['zikuy'] += zikuy

        if 'ОЎОюОЊОЎ ОЌОЋОЦ' in maslul and zikuy != 0:
            semel_data[semel]['zikuy'] += zikuy

    # ОЉОаОћ rows
    rows = []
    for semel, data in semel_data.items():
        idx_entry = welfare_index.get(semel, {})
        rows.append({
            "semel":          semel,
            "name":           data['name'],
            "debit_account":  idx_entry.get('debit', ''),
            "credit_account": idx_entry.get('credit', ''),
            "has_ОъОъОЕОюОћ":     data['has_ОъОъОЕОюОћ'],
            "debit_total":    _to_decimal(data['debit_total']),
            "zikuy_hodesh":   _to_decimal(data['zikuy']),
            "in_index":       bool(idx_entry),
        })

    total_debit  = sum(r['debit_total']       for r in rows if r['debit_account'] and r['has_ОъОъОЕОюОћ'])
    total_credit = sum(abs(r['zikuy_hodesh']) for r in rows if r['credit_account'] and r['zikuy_hodesh'] != 0)

    return {
        "municipality":    municipality,
        "period":          period_label,
        "month":           period_month,
        "year":            period_year,
        "rows":            rows,
        "total_rows":      len([r for r in rows if r['debit_total'] != 0 or r['zikuy_hodesh'] != 0]),
        "missing_index":   [r for r in rows if not r['in_index'] and (r['debit_total'] > 0 or r['zikuy_hodesh'] != 0)],
        "row_errors":      [],
        "total_debit":     float(total_debit),
        "total_credit":    float(total_credit),
        "summary_mishrad": float(summary_mishrad) if summary_mishrad else float(total_debit),
        "summary_choz":    float(summary_choz)    if summary_choz    else None,
        "balance_ok":      True,
    }


def apply_welfare_splits(parsed: dict) -> tuple:
    """
    ОюОЋОњОЎОДОћ ОАОЋОцОЎОф:
    - mishrad: ОЌОЎОЋОЉОЎРєњОЌОЋОЉОћ, ОЕОюОЎОюОЎРєњОќОЏОЋОф
    - zikuy: ОЌОЎОЋОЉОЎРєњОќОЏОЋОф, ОЕОюОЎОюОЎРєњОЌОЋОЉОћ
    - ОЌОЋ"Оќ: ОъОћОЊОЋОЌ ОЉОюОЉОЊ (summary_choz)
    """
    matched = []
    missing = []

    for row in parsed["rows"]:
        debit_total = row["debit_total"]
        zikuy       = row["zikuy_hodesh"]

        if debit_total == Decimal("0") and zikuy == Decimal("0"):
            continue

        if not row["in_index"]:
            missing.append({**row, "error": f"ОАОбОЎОБ {row['semel']} ОюОљ ОаОъОдОљ ОЉ-INDEX"})
            continue

        # mishrad: ОЌОЎОЋОЉОЎРєњОЌОЋОЉОћ, ОЕОюОЎОюОЎРєњОќОЏОЋОф
        if debit_total != Decimal("0"):
            if debit_total > Decimal("0") and row["debit_account"]:
                matched.append({
                    "semel":       row["semel"],
                    "name":        row["name"],
                    "account":     row["debit_account"],
                    "amount":      float(debit_total),
                    "side":        "debit",
                    "description": f"ОеОЋОЋОЌОћ {row['semel']} {row['name']}",
                })
            elif debit_total < Decimal("0") and row["credit_account"]:
                matched.append({
                    "semel":       row["semel"],
                    "name":        row["name"],
                    "account":     row["credit_account"],
                    "amount":      float(abs(debit_total)),
                    "side":        "credit",
                    "description": f"ОеОЋОЋОЌОћ {row['semel']} {row['name']}",
                })

        # zikuy: ОЌОЎОЋОЉОЎРєњОќОЏОЋОф, ОЕОюОЎОюОЎРєњОЌОЋОЉОћ
        if zikuy != Decimal("0"):
            if zikuy > Decimal("0") and row["credit_account"]:
                matched.append({
                    "semel":       row["semel"],
                    "name":        row["name"],
                    "account":     row["credit_account"],
                    "amount":      float(zikuy),
                    "side":        "credit",
                    "description": f"ОеОЋОЋОЌОћ {row['semel']} {row['name']}",
                })
            elif zikuy < Decimal("0") and row["debit_account"]:
                matched.append({
                    "semel":       row["semel"],
                    "name":        row["name"],
                    "account":     row["debit_account"],
                    "amount":      float(abs(zikuy)),
                    "side":        "debit",
                    "description": f"ОеОЋОЋОЌОћ {row['semel']} {row['name']}",
                })

    # ОЕОЋОеОф ОЌОЋ"Оќ Рђћ ОъОћОЊОЋОЌ ОЉОюОЉОЊ, ОюОюОљ ОљОЎОќОЋОЪ ОъОюОљОЏОЋОфОЎ
    summary_choz = parsed.get("summary_choz")
    if summary_choz and Decimal(str(summary_choz)) > Decimal("0"):
        matched.append({
            "semel":       "ОЌОЋОќ",
            "name":        'ОЌОЋ"Оќ ОъОЕОеОЊ ОћОеОЋОЋОЌОћ',
            "account":     "700000000",
            "amount":      float(Decimal(str(summary_choz))),
            "side":        "credit",
            "description": 'ОЌОЋ"Оќ ОъОЕОеОЊ ОћОеОЋОЋОЌОћ',
        })

    # ОЉОЊОЎОДОф ОфОДОЎОаОЋОф ОЉОюОЉОЊ
    summary_mishrad = parsed.get("summary_mishrad")
    if summary_mishrad and summary_choz:
        total_credit = sum(Decimal(str(r["amount"])) for r in matched if r["side"] == "credit")
        gap = abs(total_credit - Decimal(str(summary_mishrad)))
        if gap > Decimal("1"):
            print(f"[WELFARE] WARNING: credit gap = {gap} (total_credit={total_credit}, mishrad={summary_mishrad})")

    return matched, missing

# routers/journal.py – Production-ready v2
# transactions + validation + audit + settings

from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from typing import Optional
from uuid import UUID, uuid4
router = APIRouter()

def get_db():
    import db
    return db

# =============================================
# SCHEMAS
# =============================================

class JournalLine(BaseModel):
    account:     str
    description: Optional[str] = None
    debit:       float = 0.0
    credit:      float = 0.0
    reference:   Optional[str] = None
    key_value:   Optional[str] = None

class JournalEntryCreate(BaseModel):
    municipality_id: str
    template_id:     str
    period:          str
    source_file:     Optional[str] = None
    notes:           Optional[str] = None
    lines:           list[JournalLine]

class JournalEntryUpdate(BaseModel):
    status: Optional[str] = None
    notes:  Optional[str] = None
    lines:  Optional[list[JournalLine]] = None

# =============================================
# HELPERS
# =============================================

async def get_vendor_account(municipality_id: str, template: str = "bezeq") -> str:
    """שלוף חשבון ספק מ-settings לפי template"""
    row = await get_db().fetch_one(
        """SELECT value FROM municipality_settings
           WHERE municipality_id = :muni AND template_name = :tmpl AND key = 'vendor_account'""",
        values={"muni": municipality_id, "tmpl": template}
    )
    if row and row["value"]:
        return row["value"]
    # fallback לברירת מחדל לפי סוג
    default = "7000000000" if template == "electricity" else "6000203000"
    return default

async def write_audit(municipality_id: str, action: str, entity_type: str,
                      entity_id: str, before=None, after=None):
    import json
    await get_db().execute(
        """INSERT INTO audit_log (id, municipality_id, action, entity_type, entity_id, before_data, after_data)
           VALUES (:id, :muni, :action, :entity, :eid, :before, :after)""",
        values={
            "id":     str(uuid4()),
            "muni":   municipality_id,
            "action": action,
            "entity": entity_type,
            "eid":    entity_id,
            "before": json.dumps(before, ensure_ascii=False) if before else None,
            "after":  json.dumps(after,  ensure_ascii=False) if after  else None,
        }
    )

def validate_lines(lines: list[JournalLine]) -> None:
    """בדיקות תקינות על שורות הפקודה"""
    if not lines:
        raise HTTPException(status_code=422, detail="פקודה חייבת להכיל לפחות שורה אחת")

    for i, line in enumerate(lines):
        if not line.account or line.account.strip() == "":
            raise HTTPException(status_code=422, detail=f"שורה {i+1}: חסר קוד חשבון")
        if line.credit < 0:
            raise HTTPException(status_code=422, detail=f"שורה {i+1}: סכום זכות לא יכול להיות שלילי")
        if line.debit > 0 and line.credit > 0:
            raise HTTPException(status_code=422, detail=f"שורה {i+1}: לא ניתן לרשום חובה וזכות באותה שורה")

    total_debit  = round(sum(l.debit  for l in lines), 2)
    total_credit = round(sum(l.credit for l in lines), 2)
    if abs(total_debit - total_credit) > 0.10:
        raise HTTPException(
            status_code=422,
            detail=f"פקודה לא מאוזנת: חובה={total_debit} זכות={total_credit} הפרש={abs(total_debit-total_credit):.2f}"
        )

# =============================================
# POST /journal-entries
# =============================================

@router.post("", status_code=201)
async def create_journal_entry(payload: JournalEntryCreate):
    # בדיקת כפילות
    existing = await get_db().fetch_one(
        """SELECT id, reference_num FROM journal_entries
           WHERE municipality_id = :muni AND template_id = :tmpl AND period = :period""",
        values={"muni": payload.municipality_id, "tmpl": payload.template_id, "period": payload.period}
    )
    if existing:
        raise HTTPException(
            status_code=409,
            detail=f"קיימת פקודה לתקופה {payload.period}: {existing['reference_num']}"
        )

    # validation
    validate_lines(payload.lines)

    entry_id = str(uuid4())
    ref_num  = f"JRN-{payload.period.replace('-','')}-{entry_id[:6].upper()}"
    total    = round(sum(l.debit for l in payload.lines if l.debit), 2)

    # Transaction – הכל או כלום
    async with get_db().transaction():
        await get_db().execute(
            """INSERT INTO journal_entries
               (id, municipality_id, template_id, period, reference_num,
                source_file, total_amount, notes, status)
               VALUES (:id,:muni,:tmpl,:period,:ref,:src,:total,:notes,'draft')""",
            values={
                "id": entry_id, "muni": payload.municipality_id,
                "tmpl": payload.template_id, "period": payload.period,
                "ref": ref_num, "src": payload.source_file,
                "total": total, "notes": payload.notes,
            }
        )
        for i, line in enumerate(payload.lines, 1):
            await get_db().execute(
                """INSERT INTO journal_lines
                   (id,entry_id,line_num,account,description,debit,credit,reference,key_value)
                   VALUES (:id,:entry,:num,:acct,:desc,:debit,:credit,:ref,:key)""",
                values={
                    "id": str(uuid4()), "entry": entry_id, "num": i,
                    "acct": line.account, "desc": line.description,
                    "debit": line.debit, "credit": line.credit,
                    "ref": line.reference, "key": line.key_value,
                }
            )

    await write_audit(payload.municipality_id, "CREATE", "journal_entry", entry_id,
                      after={"reference_num": ref_num, "period": payload.period, "total": total})

    return {"id": entry_id, "reference_num": ref_num, "status": "draft", "total": total}

# =============================================
# GET /journal-entries/check-period
# חייב להיות לפני /{entry_id}!
# =============================================

@router.get("/check-period")
async def check_period_route(municipality_id: str, template_id: str, period: str):
    existing = await get_db().fetch_one(
        """SELECT id, reference_num FROM journal_entries
           WHERE municipality_id = :muni AND template_id = :tmpl AND period = :period""",
        values={"muni": municipality_id, "tmpl": template_id, "period": period}
    )
    if existing:
        return {"exists": True, "id": existing["id"], "reference_num": existing["reference_num"]}
    return {"exists": False}

# =============================================
# GET /journal-entries/budget-forecast
# תחזית תקציב גנרית – לכל מודול
# =============================================

@router.get("/budget-forecast")
async def budget_forecast(
    municipality_id: str,
    template_name: str,          # bezeq / electricity / welfare / cellcom
    target_year: int = 2027,
):
    """
    Generic 12-month budget forecast for any module.
    Analyzes historical journal data, detects trends, projects forward.
    """
    db = get_db()

    # Resolve template_name → template_id
    tmpl_row = await db.fetch_one(
        "SELECT id FROM templates WHERE name = :name",
        values={"name": template_name}
    )
    if not tmpl_row:
        return {"target_year": target_year, "accounts": [], "total": 0,
                "months_data": 0, "message": f"תבנית {template_name} לא נמצאה"}

    template_id = str(tmpl_row["id"])

    # Fetch all journal lines (debit > 0) for this municipality + template
    rows = await db.fetch_all(
        """
        SELECT je.period, jl.account, jl.description, jl.debit, jl.key_value
        FROM journal_lines jl
        JOIN journal_entries je ON je.id = jl.entry_id
        WHERE je.municipality_id = :muni
          AND je.template_id = :tmpl
          AND je.is_active = TRUE
          AND jl.debit > 0
        ORDER BY je.period
        """,
        values={"muni": municipality_id, "tmpl": template_id}
    )

    MODULE_LABELS = {
        "bezeq": "בזק", "electricity": "חשמל",
        "welfare": "רווחה", "cellcom": "סלקום",
    }
    label = MODULE_LABELS.get(template_name, template_name)

    if not rows:
        return {"target_year": target_year, "accounts": [], "total": 0,
                "months_data": 0, "message": f"אין נתוני {label} במערכת"}

    # Group by key_value (contract/phone/ID) → list of (period, amount, account)
    contract_data: dict[str, list] = {}
    contract_account: dict[str, str] = {}
    all_periods: set[str] = set()

    for r in rows:
        kv = r["key_value"] or ""
        if not kv:
            continue
        period = r["period"]
        amount = float(r["debit"] or 0)
        account = r["account"]
        all_periods.add(period)
        if kv not in contract_data:
            contract_data[kv] = []
        contract_data[kv].append({"period": period, "amount": amount})
        contract_account[kv] = account

    sorted_periods = sorted(all_periods)
    months_count = len(sorted_periods)

    if months_count == 0:
        return {"target_year": target_year, "accounts": [], "total": 0, "months_data": 0}

    # Per-contract forecast with trend analysis
    account_forecasts: dict[str, dict] = {}

    for kv, data_points in contract_data.items():
        account = contract_account[kv]
        amounts_by_period: dict[str, float] = {}
        for dp in data_points:
            p = dp["period"]
            amounts_by_period[p] = amounts_by_period.get(p, 0) + dp["amount"]

        sorted_amounts = [(p, amounts_by_period[p]) for p in sorted(amounts_by_period.keys())]
        n = len(sorted_amounts)
        if n == 0:
            continue

        total_amount = sum(a for _, a in sorted_amounts)
        avg_monthly = total_amount / n

        # Linear regression trend
        monthly_trend = 0.0
        if n >= 3:
            x_vals = list(range(n))
            y_vals = [a for _, a in sorted_amounts]
            x_mean = sum(x_vals) / n
            y_mean = sum(y_vals) / n
            num = sum((x - x_mean) * (y - y_mean) for x, y in zip(x_vals, y_vals))
            den = sum((x - x_mean) ** 2 for x in x_vals)
            if den > 0:
                monthly_trend = num / den
                max_trend = avg_monthly * 0.2
                monthly_trend = max(-max_trend, min(max_trend, monthly_trend))

        last_amount = sorted_amounts[-1][1] if sorted_amounts else avg_monthly
        effective_n = max(n, months_count)
        annual_base = total_amount * (12 / effective_n) if effective_n < 12 else total_amount

        month_forecasts = []
        for m in range(12):
            if n >= 6:
                projected = last_amount + monthly_trend * (m + 1)
                projected = max(projected, avg_monthly * 0.3)
            else:
                projected = avg_monthly + monthly_trend * (m - 5.5)
                projected = max(projected, avg_monthly * 0.5)
            month_forecasts.append(round(projected, 2))

        forecast_sum = sum(month_forecasts)
        if forecast_sum > 0:
            scale = annual_base / forecast_sum
            month_forecasts = [round(v * scale, 2) for v in month_forecasts]

        if account not in account_forecasts:
            account_forecasts[account] = {
                "account": account, "contracts": 0,
                "historical_total": 0, "historical_months": 0,
                "trend": "stable", "months": [0.0] * 12, "total": 0,
            }

        af = account_forecasts[account]
        af["contracts"] += 1
        af["historical_total"] += total_amount
        af["historical_months"] = max(af["historical_months"], n)

        for i in range(12):
            af["months"][i] += month_forecasts[i]
            af["months"][i] = round(af["months"][i], 2)

        if monthly_trend > avg_monthly * 0.03:
            if af["trend"] != "decline":
                af["trend"] = "growth"
        elif monthly_trend < -avg_monthly * 0.03:
            if af["trend"] != "growth":
                af["trend"] = "decline"

    # Totals + account names from indexes
    result_accounts = []
    grand_total = 0
    for acct, af in sorted(account_forecasts.items()):
        af["total"] = round(sum(af["months"]), 2)
        af["historical_total"] = round(af["historical_total"], 2)
        grand_total += af["total"]
        result_accounts.append(af)

    account_names = {}
    idx_rows = await db.fetch_all(
        """SELECT DISTINCT account_code, connection_name FROM indexes
           WHERE municipality_id = :muni AND template_id = :tmpl AND active = TRUE""",
        values={"muni": municipality_id, "tmpl": template_id}
    )
    for ir in idx_rows:
        acct = ir["account_code"]
        name = ir["connection_name"] or ""
        if acct not in account_names and name:
            account_names[acct] = name

    for af in result_accounts:
        af["account_name"] = account_names.get(af["account"], "")

    return {
        "target_year": target_year,
        "accounts": result_accounts,
        "total": round(grand_total, 2),
        "months_data": months_count,
        "contracts_count": len(contract_data),
        "periods_range": f"{sorted_periods[0]} – {sorted_periods[-1]}" if sorted_periods else "",
    }

# =============================================
# GET /journal-entries
# =============================================

@router.get("")
async def list_journal_entries(
    municipality_id: str,
    period:  Optional[str] = None,
    status:  Optional[str] = None,
    limit:   int = 50,
):
    filters = ["je.municipality_id = :muni", "je.is_active = TRUE"]
    values  = {"muni": municipality_id, "limit": limit}
    if period:
        filters.append("je.period = :period"); values["period"] = period
    if status:
        filters.append("je.status = :status"); values["status"] = status

    rows = await get_db().fetch_all(
        f"""SELECT je.id, je.reference_num, je.period, je.status,
                   je.total_amount, je.source_file, je.created_at,
                   je.template_id::text AS template_id,
                   t.name AS template_key, t.display_name AS template_name
            FROM journal_entries je
            JOIN templates t ON t.id = je.template_id
            WHERE {' AND '.join(filters)}
            ORDER BY je.created_at DESC LIMIT :limit""",
        values=values
    )
    return [dict(r) for r in rows]

# =============================================
# GET /journal-entries/{id}
# =============================================

@router.get("/{entry_id}")
async def get_journal_entry(entry_id: str):
    print(f"EXPORT ENTRY ID: {entry_id}")
    entry = await get_db().fetch_one(
        """SELECT je.*, t.display_name AS template_name, t.name AS template_key,
                  m.name AS municipality_name
           FROM journal_entries je
           JOIN templates t ON t.id = je.template_id
           JOIN municipalities m ON m.id = je.municipality_id
           WHERE je.id = :id AND je.is_active = TRUE""",
        values={"id": entry_id}
    )
    if not entry:
        raise HTTPException(status_code=404, detail="פקודה לא נמצאה")
    lines = await get_db().fetch_all(
        "SELECT * FROM journal_lines WHERE entry_id = :id ORDER BY line_num",
        values={"id": entry_id}
    )
    return {**dict(entry), "lines": [dict(l) for l in lines]}

# =============================================
# PATCH /journal-entries/{id}
# =============================================

@router.patch("/{entry_id}")
async def update_journal_entry(entry_id: str, payload: JournalEntryUpdate):
    existing = await get_db().fetch_one(
        "SELECT id, status, municipality_id FROM journal_entries WHERE id = :id",
        values={"id": entry_id}
    )
    if not existing:
        raise HTTPException(status_code=404, detail="פקודה לא נמצאה")
    if existing["status"] in ("exported", "locked"):
        raise HTTPException(status_code=400, detail=f"לא ניתן לערוך פקודה בסטטוס {existing['status']}")

    VALID_TRANSITIONS = {
        "draft":    ["ready", "draft"],
        "ready":    ["draft", "exported"],
        "exported": [],
        "locked":   [],
        "approved": ["exported"],
    }
    if payload.status and payload.status not in VALID_TRANSITIONS.get(existing["status"], []):
        raise HTTPException(
            status_code=400,
            detail=f"מעבר סטטוס לא חוקי: {existing['status']} → {payload.status}"
        )

    async with get_db().transaction():
        if payload.status:
            await get_db().execute(
                "UPDATE journal_entries SET status = :status WHERE id = :id",
                values={"status": payload.status, "id": entry_id}
            )
        if payload.lines is not None:
            validate_lines(payload.lines)
            await get_db().execute(
                "DELETE FROM journal_lines WHERE entry_id = :id", values={"id": entry_id}
            )
            for i, line in enumerate(payload.lines, 1):
                await get_db().execute(
                    """INSERT INTO journal_lines
                       (id,entry_id,line_num,account,description,debit,credit,reference,key_value)
                       VALUES (:id,:entry,:num,:acct,:desc,:debit,:credit,:ref,:key)""",
                    values={
                        "id": str(uuid4()), "entry": entry_id, "num": i,
                        "acct": line.account, "desc": line.description,
                        "debit": line.debit, "credit": line.credit,
                        "ref": line.reference, "key": line.key_value,
                    }
                )
            total = round(sum(l.debit for l in payload.lines), 2)
            await get_db().execute(
                "UPDATE journal_entries SET total_amount = :t WHERE id = :id",
                values={"t": total, "id": entry_id}
            )

    await write_audit(str(existing["municipality_id"]), "UPDATE", "journal_entry", entry_id,
                      before={"status": existing["status"]},
                      after={"status": payload.status} if payload.status else None)
    return {"id": entry_id, "updated": True}

# =============================================
# DELETE /journal-entries/{id}
# =============================================

@router.delete("/{entry_id}")
async def delete_journal_entry(entry_id: str):
    existing = await get_db().fetch_one(
        "SELECT id, status, municipality_id, reference_num FROM journal_entries WHERE id = :id",
        values={"id": entry_id}
    )
    if not existing:
        raise HTTPException(status_code=404, detail="פקודה לא נמצאה")
    if existing["status"] in ("locked",):
        raise HTTPException(status_code=400, detail="לא ניתן למחוק פקודה נעולה")

    async with get_db().transaction():
        await get_db().execute("DELETE FROM journal_lines WHERE entry_id = :id", values={"id": entry_id})
        await get_db().execute("DELETE FROM journal_entries WHERE id = :id",     values={"id": entry_id})

    await write_audit(str(existing["municipality_id"]), "DELETE", "journal_entry", entry_id,
                      before={"reference_num": existing["reference_num"], "status": existing["status"]})
    return {"deleted": True}

# =============================================
# GET /journal-entries/{id}/export
# =============================================

@router.get("/{entry_id}/export")
async def export_journal_entry(entry_id: str):
    import io
    import json
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from fastapi.responses import StreamingResponse

    entry = await get_journal_entry(entry_id)

    period = entry["period"]
    y, m   = int(period.split("-")[0]), int(period.split("-")[1])
    date_str = f"1/{m}/{y}"
    year_num = y

    extra = {}
    try:
        if entry.get("notes"):
            extra = json.loads(entry["notes"])
    except Exception:
        pass
    invoice_num    = extra.get("invoice_num", "")
    date_from      = extra.get("date_from", "")
    date_to        = extra.get("date_to", "")
    billing_period = f"{date_from}-{date_to}" if date_from and date_to else ""

    # זיהוי סוג הפקודה
    template_key = (
        entry.get("template_key") or
        entry.get("template_name_key") or
        entry.get("source_type") or
        "bezeq"
    )
    is_electricity = template_key == "electricity"
    is_welfare     = template_key == "welfare"

    vendor_account = await get_vendor_account(entry["municipality_id"], template_key)
    # עדכן גם את billing_period לחשמל
    if is_electricity and not billing_period:
        billing_period = extra.get("date_from", "") + ("-" + extra.get("date_to", "") if extra.get("date_to") else "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "פקודת יומן"
    ws.sheet_view.rightToLeft = True

    headers = ["שנה","מספר פקודה","סטטוס","מקור","תאריך רישום","תאריך ערך",
               "מספר חשבון","מספר חשבון נגדי","אסמכתא 1","אסמכתא 2",
               "חשבונית","תאריך חשבון","פרטים","ח/ז","קוד מיון",'סה"כ']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    async def get_conn_name(key_value: str) -> str:
        if key_value:
            idx_row = await get_db().fetch_one(
                """SELECT connection_name FROM indexes
                   WHERE municipality_id = :muni AND key_value = :key AND active = TRUE LIMIT 1""",
                values={"muni": entry["municipality_id"], "key": key_value}
            )
            if idx_row and idx_row["connection_name"]:
                return idx_row["connection_name"]
        return entry["municipality_name"] or ""

    is_welfare = template_key == "welfare"

    if is_welfare:
        # רווחה: כל שורה (חובה וזכות) מיוצאת בנפרד
        for line in entry["lines"]:
            debit_amt  = round(float(line["debit"]  or 0), 2)
            credit_amt = round(float(line["credit"] or 0), 2)
            if debit_amt == 0 and credit_amt == 0:
                continue
            amount = debit_amt if debit_amt > 0 else credit_amt
            side   = 1 if debit_amt > 0 else 2
            ref    = (line["key_value"] or "").strip()
            desc   = (line.get("description") or ref or "רווחה").strip()
            ws.append([
                year_num, "", 10, 6, date_str, date_str,
                line["account"], "",
                ref, "", "", "",
                desc, side, "",
                amount,
            ])
    else:
        debit_lines = [l for l in entry["lines"] if float(l["debit"] or 0) > 0]
        for line in debit_lines:
            subscriber = (line["key_value"] or "").strip()
            if is_electricity:
                conn_name = await get_conn_name(subscriber)
                if not conn_name or conn_name == entry.get("municipality_name", ""):
                    conn_name = (line.get("description") or subscriber)
            else:
                conn_name = await get_conn_name(subscriber)
            details = f"{conn_name} {billing_period}".strip()
            ws.append([
                year_num, "", 10, 6, date_str, date_str,
                line["account"], "",
                subscriber, invoice_num, "", "",
                details, 1, "",
                round(float(line["debit"]), 2),
            ])

        total = round(sum(float(l["debit"] or 0) for l in debit_lines), 2)
        credit_label = "חשמל - חשבון חודשי" if is_electricity else "בזק - חשבון חודשי"
        ws.append([year_num, "", 10, 6, date_str, date_str,
                   vendor_account, "", "", "", "", "",
                   credit_label, 2, "", total])

    col_widths = [8,14,8,8,14,14,16,16,14,12,12,14,40,6,10,12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # עדכן סטטוס ל-exported
    await get_db().execute(
        "UPDATE journal_entries SET status = 'exported' WHERE id = :id AND status IN ('ready', 'draft')",
        values={"id": entry_id}
    )
    await write_audit(str(entry["municipality_id"]), "EXPORT", "journal_entry", entry_id,
                      after={"reference_num": entry["reference_num"]})

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"journal_{entry['reference_num']}_{period}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

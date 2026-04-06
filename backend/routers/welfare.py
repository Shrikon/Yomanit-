# routers/welfare.py – upload + preview + approve לרווחה
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from parsers.welfare import parse_welfare, apply_welfare_splits, WelfareParserError
from decimal import Decimal, ROUND_HALF_UP
from uuid import uuid4
import json
import time
from index_cache import get_index as get_cached_index

router = APIRouter()

WELFARE_TEMPLATE_NAME = "welfare"


# ─────────────────────────────────────────────
# POST /upload/welfare – preview
# ─────────────────────────────────────────────
@router.post("")
async def upload_welfare(
    file:            UploadFile = File(...),
    municipality_id: str        = Form(...),
    month:           int        = Form(None),
):
    t0 = time.time()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="קובץ ריק")
    print(f"[PERF] READ FILE: {time.time()-t0:.3f}s  size={len(content)} bytes")

    # טען אינדקס מה-Cache (קריאה אחת ל-DB בלבד לךת אחר cached)
    from main import database as _db
    try:
        t1 = time.time()
        welfare_tmpl = await _db.fetch_one(
            "SELECT id FROM templates WHERE name = 'welfare'", values={}
        )
        welfare_tmpl_id = str(welfare_tmpl["id"]) if welfare_tmpl else None

        index_map = {}
        if welfare_tmpl_id:
            index_map = await get_cached_index(municipality_id, welfare_tmpl_id, _db)
        print(f"[PERF] INDEX (cached): {time.time()-t1:.3f}s  entries={len(index_map)}")

        if index_map:
            print(f"[INDEX] SOURCE=DB  entries={len(index_map)}")
            for semel_check in ["721060", "722020", "242410"]:
                val = index_map.get(semel_check)
                print(f"[INDEX] semel={semel_check} → {val}")
        else:
            print(f"[INDEX] SOURCE=STATIC (DB index_map empty!)")

        use_index = index_map if index_map else None

    except Exception as e:
        print(f"[INDEX] ERROR loading index: {e}")
        use_index = None

    try:
        t3 = time.time()
        parsed = parse_welfare(content, month=month, index_map=use_index)
        print(f"[PERF] PARSE_WELFARE: {time.time()-t3:.3f}s  rows={parsed.get('total_rows')}")
    except WelfareParserError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"שגיאת פרסור: {str(e)}")

    t4 = time.time()
    matched, missing = apply_welfare_splits(parsed)
    print(f"[PERF] APPLY_SPLITS: {time.time()-t4:.3f}s  matched={len(matched)} missing={len(missing)}")
    print(f"[PERF] TOTAL: {time.time()-t0:.3f}s")

    # בנה שורות תצוגה
    rows_out = []
    for r in matched:
        rows_out.append({
            "semel":       r["semel"],
            "name":        r["name"],
            "account":     r["account"],
            "amount":      r["amount"],
            "side":        r["side"],
            "description": r["description"],
            "status":      "ok",
        })
    for r in missing:
        amt = float(r.get("chiuv_hodesh", 0) or r.get("mishrad", 0))
        rows_out.append({
            "semel":       r["semel"],
            "name":        r["name"],
            "account":     "",
            "amount":      amt,
            "side":        "debit" if amt >= 0 else "credit",
            "description": r["name"],
            "status":      "missing_index",
            "error":       r.get("error", ""),
        })

    # Missing index entries no longer block approval
    can_approve = True

    return {
        "filename":      file.filename,
        "template":      "welfare",
        "municipality":  parsed["municipality"],
        "period":        parsed["period"],
        "month":         parsed["month"],
        "year":          parsed.get("year", 2026),
        "total_rows":    len(rows_out),
        "matched":       len(matched),
        "missing":       len(missing),
        "rows":          rows_out,
        "total_debit":   parsed["total_debit"],
        "total_credit":  parsed["total_credit"],
        "can_approve":   can_approve,
        "missing_index": parsed["missing_index"],
    }


# ─────────────────────────────────────────────
# POST /upload/welfare/approve – יצירת פקודה
# ─────────────────────────────────────────────
from pydantic import BaseModel
from typing import List, Optional

class WelfareLineIn(BaseModel):
    semel:       str
    account:     str
    amount:      float
    side:        str   # debit | credit
    description: Optional[str] = None

from datetime import datetime as _dt

class WelfareApproveIn(BaseModel):
    municipality_id: str
    period:          str
    month:           int
    year:            int = _dt.now().year
    source_file:     Optional[str] = None
    source_file_b64: Optional[str] = None  # base64-encoded Excel for treasurer report
    lines:           List[WelfareLineIn]


@router.post("/approve", status_code=201)
async def approve_welfare(payload: WelfareApproveIn):
    from main import database

    if not payload.lines:
        raise HTTPException(status_code=422, detail="אין שורות לאישור")

    # שלוף template_id
    tmpl = await database.fetch_one(
        "SELECT id FROM templates WHERE name = 'welfare'", values={}
    )
    if not tmpl:
        tmpl_id = str(uuid4())
        await database.execute(
            """INSERT INTO templates (id, name, display_name)
               VALUES (:id, 'welfare', 'רווחה')
               ON CONFLICT (name) DO NOTHING""",
            values={"id": tmpl_id}
        )
        tmpl = await database.fetch_one(
            "SELECT id FROM templates WHERE name = 'welfare'", values={}
        )

    template_id = str(tmpl["id"])
    period_str  = f"{payload.year}-{payload.month:02d}"

    # בדיקת כפילות
    existing = await database.fetch_one(
        """SELECT id, reference_num FROM journal_entries
           WHERE municipality_id = :muni AND template_id = :tmpl AND period = :period
             AND is_active = TRUE""",
        values={"muni": payload.municipality_id, "tmpl": template_id, "period": period_str}
    )
    if existing:
        raise HTTPException(status_code=409,
            detail=f"קיימת פקודה לתקופה {period_str}: {existing['reference_num']}")

    # חשב סכומים
    total_debit  = Decimal("0")
    total_credit = Decimal("0")
    for line in payload.lines:
        amt = Decimal(str(line.amount)).quantize(Decimal("0.01"), ROUND_HALF_UP)
        if line.side == "debit":
            total_debit += amt
        else:
            total_credit += amt

    # שלוף חשבון חו"ז משרד הרווחה
    ministry_setting = await database.fetch_one(
        """SELECT value FROM municipality_settings
           WHERE municipality_id = :muni AND template_name = 'welfare' AND key = 'ministry_account'""",
        values={"muni": payload.municipality_id}
    )
    ministry_account = ministry_setting["value"] if ministry_setting else "700000000"

    # חשב הפרש לחו"ז
    diff = (total_credit - total_debit).quantize(Decimal("0.01"), ROUND_HALF_UP)

    entry_id = str(uuid4())
    ref_num  = f"WLF-{period_str.replace('-','')}-{entry_id[:6].upper()}"

    async with database.transaction():
        await database.execute(
            """INSERT INTO journal_entries
               (id, municipality_id, template_id, period, reference_num,
                source_file, total_amount, status, source_type,
                source_period_month, source_period_year)
               VALUES (:id,:muni,:tmpl,:period,:ref,:src,:total,'draft','welfare',:month,:year)""",
            values={
                "id":    entry_id,
                "muni":  payload.municipality_id,
                "tmpl":  template_id,
                "period": period_str,
                "ref":   ref_num,
                "src":   payload.source_file,
                "total": float(total_debit),
                "month": payload.month,
                "year":  payload.year,
            }
        )

        # שורות פקודה
        for i, line in enumerate(payload.lines, 1):
            amt = Decimal(str(line.amount)).quantize(Decimal("0.01"), ROUND_HALF_UP)
            await database.execute(
                """INSERT INTO journal_lines
                   (id,entry_id,line_num,account,description,debit,credit,reference,key_value)
                   VALUES (:id,:entry,:num,:acct,:desc,:debit,:credit,:ref,:key)""",
                values={
                    "id":     str(uuid4()),
                    "entry":  entry_id,
                    "num":    i,
                    "acct":   line.account,
                    "desc":   line.description or f"רווחה {line.semel}",
                    "debit":  float(amt) if line.side == "debit" else 0.0,
                    "credit": float(amt) if line.side == "credit" else 0.0,
                    "ref":    line.semel,
                    "key":    line.semel,
                }
            )

        # שורת איזון — חו"ז משרד הרווחה
        if diff != 0:
            line_num = len(payload.lines) + 1
            await database.execute(
                """INSERT INTO journal_lines
                   (id,entry_id,line_num,account,description,debit,credit,reference,key_value)
                   VALUES (:id,:entry,:num,:acct,:desc,:debit,:credit,:ref,:key)""",
                values={
                    "id":     str(uuid4()),
                    "entry":  entry_id,
                    "num":    line_num,
                    "acct":   ministry_account,
                    "desc":   f"חו\"ז משרד הרווחה {period_str}",
                    "debit":  float(abs(diff)) if diff > 0 else 0.0,
                    "credit": float(abs(diff)) if diff < 0 else 0.0,
                    "ref":    "",
                    "key":    "",
                }
            )

        # בדוק אם כל השורות יש להן חשבון — אם כן, קבע אוטומטית כ-approved
        all_lines_have_account = all(
            (line.account or '').strip() != ''
            for line in payload.lines
        )
        final_status = 'approved' if all_lines_have_account else 'draft'
        if final_status == 'approved':
            await database.execute(
                """UPDATE journal_entries
                   SET status = 'approved', approved_at = NOW()
                   WHERE id = :id""",
                values={"id": entry_id}
            )

    # ── Update index names + auto-insert missing semels ──
    try:
        from index_cache import invalidate
        dirty = False
        # Update connection_name for all lines that have a description
        for line in payload.lines:
            desc_name = (line.description or '').strip()
            if desc_name and line.semel:
                await database.execute(
                    """UPDATE indexes SET connection_name = :name, updated_at = NOW()
                       WHERE municipality_id = :muni AND template_id = :tmpl
                         AND key_value = :key AND (connection_name IS NULL OR connection_name = '')""",
                    values={"name": desc_name, "muni": payload.municipality_id, "tmpl": template_id, "key": line.semel}
                )
                dirty = True
        # Auto-insert missing semels
        missing_semels = [
            line for line in payload.lines
            if not (line.account or '').strip()
        ]
        if missing_semels:
            for line in missing_semels:
                conn_name = (line.description or '').strip() or None
                # Insert debit placeholder
                await database.execute(
                    """INSERT INTO indexes (id, municipality_id, template_id, key_value, account_code, description, connection_name, active)
                       VALUES (:id, :muni, :tmpl, :key, '', 'debit', :conn, TRUE)
                       ON CONFLICT (municipality_id, template_id, key_value, description)
                       DO NOTHING""",
                    values={"id": str(uuid4()), "muni": payload.municipality_id, "tmpl": template_id, "key": line.semel, "conn": conn_name}
                )
                # Insert credit placeholder
                await database.execute(
                    """INSERT INTO indexes (id, municipality_id, template_id, key_value, account_code, description, connection_name, active)
                       VALUES (:id, :muni, :tmpl, :key, '', 'credit', :conn, TRUE)
                       ON CONFLICT (municipality_id, template_id, key_value, description)
                       DO NOTHING""",
                    values={"id": str(uuid4()), "muni": payload.municipality_id, "tmpl": template_id, "key": line.semel, "conn": conn_name}
                )
            dirty = True
            print(f"[WELFARE] Auto-inserted {len(missing_semels)} missing semels into indexes", flush=True)
        if dirty:
            invalidate(payload.municipality_id, template_id)
    except Exception as e:
        print(f"[WELFARE] Auto-insert index failed (non-fatal): {e}", flush=True)

    # ── Send treasurer report email (async, non-blocking) ──
    try:
        if payload.source_file_b64:
            import base64
            from email_service import send_treasurer_report
            excel_bytes = base64.b64decode(payload.source_file_b64)
            await send_treasurer_report(
                municipality_id=payload.municipality_id,
                excel_bytes=excel_bytes,
                month=payload.month,
                year=payload.year,
            )
    except Exception as e:
        print(f"[WELFARE] Email send failed (non-fatal): {e}", flush=True)

    return {
        "id":            entry_id,
        "reference_num": ref_num,
        "status":        final_status,
        "total_debit":   float(total_debit + (diff if diff > 0 else 0)),
        "total_credit":  float(total_credit + (abs(diff) if diff < 0 else 0)),
        "ministry_diff": float(diff),
        "ministry_account": ministry_account,
        "lines_count":   len(payload.lines) + (1 if diff != 0 else 0),
    }


# ─────────────────────────────────────────────
# DELETE /upload/welfare/{entry_id} – מחיקת פקודה
# ─────────────────────────────────────────────
@router.delete("/{entry_id}", status_code=200)
async def delete_welfare_entry(entry_id: str):
    from main import database

    entry = await database.fetch_one(
        """SELECT id, reference_num, status, template_id
           FROM journal_entries
           WHERE id = :id AND is_active = TRUE""",
        values={"id": entry_id}
    )
    if not entry:
        raise HTTPException(status_code=404, detail="פקודה לא נמצאה")

    # ווידא שזו פקודת רווחה
    tmpl = await database.fetch_one(
        "SELECT name FROM templates WHERE id = :id",
        values={"id": str(entry["template_id"])}
    )
    if not tmpl or tmpl["name"] != "welfare":
        raise HTTPException(status_code=400, detail="פקודה זו אינה פקודת רווחה")

    if entry["status"] == "exported":
        raise HTTPException(status_code=400, detail="לא ניתן למחוק פקודה שיוצאה")

    async with database.transaction():
        await database.execute(
            "DELETE FROM journal_lines WHERE entry_id = :id",
            values={"id": entry_id}
        )
        await database.execute(
            "UPDATE journal_entries SET is_active = FALSE WHERE id = :id",
            values={"id": entry_id}
        )

    return {"deleted": True, "reference_num": entry["reference_num"]}


# ─────────────────────────────────────────────
# POST /upload/welfare/treasurer-report
# מקבל קובץ Excel רווחה, מחזיר PDF דוח לגזבר
# ─────────────────────────────────────────────
@router.post("/treasurer-report")
async def welfare_treasurer_report(
    file: UploadFile = File(...),
):
    from fastapi.responses import Response
    from welfare_report_analyzer import parse_excel, generate_pdf
    import tempfile, os, urllib.parse

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="קובץ ריק")

    # Write to temp file for openpyxl (needs file path)
    tmp_excel = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp_pdf = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
    try:
        tmp_excel.write(content)
        tmp_excel.close()
        tmp_pdf.close()

        report = parse_excel(tmp_excel.name)
        generate_pdf(report, tmp_pdf.name)

        with open(tmp_pdf.name, "rb") as f:
            pdf_bytes = f.read()

        safe_muni = report.municipality.replace(" ", "_")
        safe_month = report.month
        filename_utf8 = urllib.parse.quote(
            f"דוח_גזבר_{safe_muni}_{safe_month}.pdf"
        )

        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition":
                    f'attachment; filename="treasurer_report.pdf"; '
                    f"filename*=UTF-8''{filename_utf8}"
            },
        )
    except WelfareParserError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"שגיאה בהפקת דוח: {str(e)}")
    finally:
        os.unlink(tmp_excel.name)
        os.unlink(tmp_pdf.name)


# ─────────────────────────────────────────────
# POST /upload/welfare/missing-report
# מחזיר Excel עם סעיפים חסרים — מקבל JSON
# ─────────────────────────────────────────────
from pydantic import BaseModel as _MBM
from typing import List as _MList

class MissingReportIn(_MBM):
    municipality_id: str
    period:          str = ""
    missing:         _MList[dict] = []

@router.post("/missing-report")
async def welfare_missing_report(payload: MissingReportIn):
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io as _io

    missing = payload.missing
    if not missing:
        raise HTTPException(status_code=204, detail="אין סעיפים חסרים")

    municipality = payload.municipality_id
    period       = payload.period

    # שלוף שם רשות מה-DB
    from main import database
    muni_row = await database.fetch_one(
        "SELECT name FROM municipalities WHERE id = :id",
        values={"id": municipality}
    )
    muni_name = muni_row["name"] if muni_row else municipality

    wb = Workbook()
    ws = wb.active
    ws.title = "סעיפים חסרים"
    ws.sheet_view.rightToLeft = True

    HEADER_BG = "1F4E79"
    WARN_BG   = "FFF2CC"
    ROW_ALT   = "DEEAF1"
    thin      = Side(style="thin", color="AAAAAA")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # שורה 1: כותרת ראשית
    NUM_COLS = 7
    ws.merge_cells(f"A1:{get_column_letter(NUM_COLS)}1")
    ws["A1"] = f"סעיפים חסרים באינדקס — {muni_name} | תקופה: {period}"
    ws["A1"].font      = Font(bold=True, size=13, color="1F4E79")
    ws["A1"].fill      = PatternFill("solid", fgColor=WARN_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # שורה 2: הנחייה
    ws.merge_cells(f"A2:{get_column_letter(NUM_COLS)}2")
    ws["A2"] = "יש להעביר רשימה זו למחלקת הרווחה לצורך קביעת חשבון חיוב וחשבון זכות לכל סעיף"
    ws["A2"].font      = Font(italic=True, size=10, color="595959")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # שורה 3: כותרות עמודות
    headers = ["סמל סעיף", "שם סעיף", "חיוב ממשלה (T)", "סעיף הוצאה", "הכנסה (K)", "סעיף הכנסה", "הערות"]
    widths  = [14, 36, 18, 16, 18, 16, 20]
    ws.row_dimensions[3].height = 20
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w

    # שורות נתונים
    for i, r in enumerate(missing):
        row  = 4 + i
        fill = PatternFill("solid", fgColor=ROW_ALT if i % 2 == 0 else "FFFFFF")
        govt   = float(r.get("govt_amount",   0) or 0)
        source = float(r.get("source_amount", 0) or 0)
        # [semel, name, T, סעיף הוצאה (ריק), K, סעיף הכנסה (ריק), הערות (ריק)]
        vals = [r.get("semel",""), r.get("name",""),
                govt   if govt   != 0 else "", "",
                source if source != 0 else "", "", ""]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill   = fill
            cell.border = border
            cell.font   = Font(size=10)
            cell.alignment = Alignment(horizontal="right")
            if col in (3, 5) and val != "":
                cell.number_format = "#,##0"

    last = 4 + len(missing)
    ws.cell(row=last, column=1, value=f'סה"כ סעיפים חסרים:').font = Font(bold=True, size=10)
    ws.cell(row=last, column=2, value=len(missing)).font = Font(bold=True, size=10)

    buf = _io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    import urllib.parse
    filename_ascii = f"welfare_missing_{period.replace('/', '-')}.xlsx"
    filename_utf8  = urllib.parse.quote(f"welfare_missing_{muni_name}_{period.replace('/', '-')}.xlsx")
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename_ascii}"; filename*=UTF-8\'\'{filename_utf8}'},
    )